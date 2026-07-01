#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
预测结果导出服务
支持导出为 Excel 和 PDF 格式
"""

import io
import sqlite3
from datetime import datetime
from typing import Dict, List, Optional, Any
from backend.config.config import Config


# ============================================================
# Excel 导出
# ============================================================

def export_prediction_excel(record: Dict[str, Any], k1_algorithms: List[Dict] = None, q1_algorithms: List[Dict] = None) -> bytes:
    """
    将单条预测记录导出为格式精美的 Excel 文件
    
    设计特色：
    - 渐变色表头
    - 置信度颜色标识（高=绿，中=黄，低=红）
    - Emoji 图标装饰
    - K1/Q1 Top5 算法明细表格
    - 品牌配色
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "预测结果报告"
    
    # ============ 颜色定义 ============
    PRIMARY = "667EEA"      # 主色调（紫蓝）
    PRIMARY_LIGHT = "E8ECFF" # 浅底色
    ACCENT_GREEN = "27AE60"  # 高置信/成功
    ACCENT_YELLOW = "F39C12" # 中置信/警告
    ACCENT_RED = "E74C3C"    # 低置信/危险
    ACCENT_BLUE = "3498DB"   # 辅助蓝
    ACCENT_PURPLE = "9B59B6" # 辅助紫
    TEXT_DARK = "2C3E50"     # 深色文字
    TEXT_GRAY = "7F8C8D"     # 灰色文字
    WHITE = "FFFFFF"
    BG_LIGHT = "F8F9FB"      # 浅灰背景
    
    # ============ 样式定义 ============
    thin_border = Border(
        left=Side(style='thin', color='D0D7E2'),
        right=Side(style='thin', color='D0D7E2'),
        top=Side(style='thin', color='D0D7E2'),
        bottom=Side(style='thin', color='D0D7E2'),
    )
    
    # 标题栏样式（主色渐变效果）
    title_font = Font(name='Arial', size=20, bold=True, color=WHITE)
    title_fill = PatternFill(start_color=PRIMARY, end_color="764BA2", fill_type='solid')
    title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 表头样式
    header_font = Font(name='Arial', size=12, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 数据行样式
    label_font = Font(name='Arial', size=11, bold=True, color=TEXT_DARK)
    label_fill = PatternFill(start_color=BG_LIGHT, end_color=BG_LIGHT, fill_type='solid')
    value_font = Font(name='Arial', size=11, color=TEXT_DARK)
    value_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 置信度颜色
    def get_confidence_fill(confidence):
        if confidence is None:
            return PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        if confidence >= 0.7:
            return PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        if confidence >= 0.5:
            return PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        return PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    
    def get_confidence_font(confidence):
        if confidence is None:
            return Font(name='Arial', size=11, color=TEXT_GRAY)
        if confidence >= 0.7:
            return Font(name='Arial', size=11, bold=True, color=ACCENT_GREEN)
        if confidence >= 0.5:
            return Font(name='Arial', size=11, bold=True, color=ACCENT_YELLOW)
        return Font(name='Arial', size=11, bold=True, color=ACCENT_RED)
    
    def get_confidence_text(confidence):
        if confidence is None:
            return "— 未评估 —"
        if confidence >= 0.8:
            return "⭐ 极高置信度"
        if confidence >= 0.7:
            return "✅ 高置信度"
        if confidence >= 0.5:
            return "⚠️ 中等置信度"
        return "🔴 低置信度"
    
    # ============ 列宽 ============
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 22
    
    # ============ 第 1-2 行：标题栏 ============
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 55
    title_cell = ws['A1']
    title_cell.value = "🎯 智能预测报告\n工程开标数据智能分析平台 V6.0"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_align
    
    # ============ 第 3 行：导出时间 ============
    ws.merge_cells('A3:C3')
    ws.row_dimensions[3].height = 28
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    time_cell = ws['A3']
    time_cell.value = f"📅 导出时间：{export_time}"
    time_cell.font = Font(name='Arial', size=10, color=TEXT_GRAY, italic=True)
    time_cell.alignment = Alignment(horizontal='center', vertical='center')
    time_cell.fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
    
    # ============ 第 5 行起：项目基本信息 ============
    row = 5
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 32
    section_cell = ws[f'A{row}']
    section_cell.value = "📋 项目基本信息"
    section_cell.font = Font(name='Arial', size=14, bold=True, color=PRIMARY)
    section_cell.fill = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type='solid')
    section_cell.border = Border(bottom=Side(style='medium', color=PRIMARY))
    section_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 项目信息字段
    info_fields = [
        ("📝 项目名称", record.get('project_name') or '— 未命名 —'),
        ("📍 开标地点", record.get('location_filter') or '不限'),
        ("📂 方法类别", f"方法{record.get('method_filter') or '不限'}"),
        ("📅 日期范围", f"{record.get('date_from') or '—'} ~ {record.get('date_to') or '—'}"),
        ("📊 数据条数", f"{record.get('data_count', 0)} 条"),
        ("🤖 AI 辅助", "✅ 是" if record.get('used_ai') else "❌ 否"),
        ("⏰ 预测时间", record.get('prediction_time') or '—'),
    ]
    
    row += 1
    for label, value in info_fields:
        ws.row_dimensions[row].height = 28
        
        # 标签列
        ws.merge_cells(f'A{row}:A{row}')
        label_cell = ws[f'A{row}']
        label_cell.value = label
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.border = thin_border
        label_cell.alignment = center_align
        
        # 值列
        ws.merge_cells(f'B{row}:C{row}')
        val_cell = ws[f'B{row}']
        val_cell.value = value
        val_cell.font = value_font
        val_cell.fill = value_fill
        val_cell.border = thin_border
        val_cell.alignment = left_align
        
        row += 1
    
    # ============ 预测结果摘要 ============
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 32
    section_cell = ws[f'A{row}']
    section_cell.value = "🎯 预测结果摘要"
    section_cell.font = Font(name='Arial', size=14, bold=True, color=PRIMARY)
    section_cell.fill = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type='solid')
    section_cell.border = Border(bottom=Side(style='medium', color=PRIMARY))
    section_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 表头
    row += 1
    ws.row_dimensions[row].height = 35
    for col_idx, header in enumerate(["📊 预测参数", "🔮 预测值", "📈 置信度"], 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_align
    
    # 预测数据行
    predictions = [
        ("方法类别", record.get('method_prediction'), record.get('method_confidence'), "方法类别预测"),
        ("K1 值", record.get('k1_prediction'), record.get('k1_confidence'), "K1 预测方法：" + (record.get('k1_method') or '—')),
        ("Q1 值", record.get('q1_prediction'), record.get('q1_confidence'), "Q1 预测方法：" + (record.get('q1_method') or '—')),
    ]
    
    row += 1
    for param, value, confidence, method_info in predictions:
        ws.row_dimensions[row].height = 35
        
        # 参数名
        c1 = ws.cell(row=row, column=1)
        c1.value = param
        c1.font = Font(name='Arial', size=11, bold=True, color=TEXT_DARK)
        c1.fill = label_fill
        c1.border = thin_border
        c1.alignment = center_align
        
        # 预测值
        c2 = ws.cell(row=row, column=2)
        c2.value = value or '—'
        c2.font = Font(name='Arial', size=14, bold=True, color=PRIMARY)
        c2.fill = value_fill
        c2.border = thin_border
        c2.alignment = center_align
        
        # 置信度
        c3 = ws.cell(row=row, column=3)
        c3.value = get_confidence_text(confidence)
        c3.font = get_confidence_font(confidence)
        c3.fill = get_confidence_fill(confidence)
        c3.border = thin_border
        c3.alignment = center_align
        
        row += 1
    
    # ============ K1 Top5 算法明细（如果有） ============
    if k1_algorithms and len(k1_algorithms) > 0:
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 32
        section_cell = ws[f'A{row}']
        section_cell.value = "🎯 K1 Top5 算法预测结果"
        section_cell.font = Font(name='Arial', size=14, bold=True, color=PRIMARY)
        section_cell.fill = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type='solid')
        section_cell.border = Border(bottom=Side(style='medium', color=PRIMARY))
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 表头
        row += 1
        ws.row_dimensions[row].height = 35
        headers = ["🔢 排名", "📐 算法名称", "🔮 预测值（置信度）"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type='solid')
            cell.border = thin_border
            cell.alignment = header_align
        
        row += 1
        for idx, algo in enumerate(k1_algorithms[:5], 1):
            ws.row_dimensions[row].height = 28
            alt_fill = PatternFill(start_color=BG_LIGHT if idx % 2 == 1 else WHITE,
                                   end_color=BG_LIGHT if idx % 2 == 1 else WHITE,
                                   fill_type='solid')
            
            c1 = ws.cell(row=row, column=1)
            c1.value = f"#{idx}"
            c1.font = Font(name='Arial', size=11, bold=True, color=PRIMARY)
            c1.fill = alt_fill
            c1.border = thin_border
            c1.alignment = center_align
            
            c2 = ws.cell(row=row, column=2)
            c2.value = algo.get('algorithm_name', '—')
            c2.font = Font(name='Arial', size=10, color=TEXT_DARK)
            c2.fill = alt_fill
            c2.border = thin_border
            c2.alignment = left_align
            
            c3 = ws.cell(row=row, column=3)
            pred_value = algo.get('prediction_value', '—')
            confidence = algo.get('confidence', 0)
            c3.value = f"{pred_value} ({confidence:.1%})"
            c3.font = Font(name='Arial', size=10, bold=True, color=PRIMARY)
            c3.fill = alt_fill
            c3.border = thin_border
            c3.alignment = center_align
            
            row += 1
    
    # ============ Q1 Top5 算法明细（如果有且非方法1） ============
    if q1_algorithms and len(q1_algorithms) > 0 and record.get('method_prediction') != '1':
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 32
        section_cell = ws[f'A{row}']
        section_cell.value = "⚙️ Q1 Top5 算法预测结果"
        section_cell.font = Font(name='Arial', size=14, bold=True, color=PRIMARY)
        section_cell.fill = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type='solid')
        section_cell.border = Border(bottom=Side(style='medium', color=PRIMARY))
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 表头
        row += 1
        ws.row_dimensions[row].height = 35
        headers = ["🔢 排名", "📐 算法名称", "🔮 预测值（置信度）"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type='solid')
            cell.border = thin_border
            cell.alignment = header_align
        
        row += 1
        for idx, algo in enumerate(q1_algorithms[:5], 1):
            ws.row_dimensions[row].height = 28
            alt_fill = PatternFill(start_color=BG_LIGHT if idx % 2 == 1 else WHITE,
                                   end_color=BG_LIGHT if idx % 2 == 1 else WHITE,
                                   fill_type='solid')
            
            c1 = ws.cell(row=row, column=1)
            c1.value = f"#{idx}"
            c1.font = Font(name='Arial', size=11, bold=True, color=PRIMARY)
            c1.fill = alt_fill
            c1.border = thin_border
            c1.alignment = center_align
            
            c2 = ws.cell(row=row, column=2)
            c2.value = algo.get('algorithm_name', '—')
            c2.font = Font(name='Arial', size=10, color=TEXT_DARK)
            c2.fill = alt_fill
            c2.border = thin_border
            c2.alignment = left_align
            
            c3 = ws.cell(row=row, column=3)
            pred_value = algo.get('prediction_value', '—')
            confidence = algo.get('confidence', 0)
            c3.value = f"{pred_value} ({confidence:.1%})"
            c3.font = Font(name='Arial', size=10, bold=True, color=PRIMARY)
            c3.fill = alt_fill
            c3.border = thin_border
            c3.alignment = center_align
            
            row += 1
    
    # ============ 底部说明 ============
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 24
    footer_cell = ws[f'A{row}']
    footer_cell.value = "⚡ 本报告由工程开标数据智能分析平台 V6.0 自动生成"
    footer_cell.font = Font(name='Arial', size=9, color=TEXT_GRAY, italic=True)
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')
    footer_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 20
    disclaimer_cell = ws[f'A{row}']
    disclaimer_cell.value = "📌 预测结果仅供参考，不构成任何决策建议"
    disclaimer_cell.font = Font(name='Arial', size=8, color='AAAAAA', italic=True)
    disclaimer_cell.alignment = Alignment(horizontal='center', vertical='center')
    disclaimer_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    # ============ 打印设置 ============
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_format.outlineLevelCol = 1
    
    # ============ 输出 ============
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# PDF 导出
# ============================================================

def export_prediction_pdf(record: Dict[str, Any], k1_algorithms: List[Dict] = None, q1_algorithms: List[Dict] = None) -> bytes:
    """
    将单条预测记录导出为精美的 PDF 报告
    
    使用 reportlab + CID 中文字体（STSong-Light）
    包含 K1 和 Q1 的 Top5 算法预测结果
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.lib.colors import (
            HexColor, white, black, Color, PCMYKColor
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            KeepTogether, PageBreak, HRFlowable
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.platypus.flowables import Flowable
    except ImportError:
        raise ImportError("需要安装 reportlab: pip install reportlab")
    
    # 注册中文字体
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    
    # ============ 颜色定义 ============
    PRIMARY = HexColor('#667EEA')
    PRIMARY_DARK = HexColor('#764BA2')
    ACCENT_GREEN = HexColor('#27AE60')
    ACCENT_YELLOW = HexColor('#F39C12')
    ACCENT_RED = HexColor('#E74C3C')
    TEXT_DARK = HexColor('#2C3E50')
    TEXT_GRAY = HexColor('#7F8C8D')
    BG_LIGHT = HexColor('#F8F9FB')
    BORDER_COLOR = HexColor('#D0D7E2')
    CONF_HIGH_BG = HexColor('#D4EDDA')
    CONF_MED_BG = HexColor('#FFF3CD')
    CONF_LOW_BG = HexColor('#F8D7DA')
    CONF_NONE_BG = HexColor('#F0F0F0')
    
    # ============ 页面设置 ============
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=25 * mm,
        bottomMargin=25 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        title="智能预测报告",
        author="工程开标数据智能分析平台 V6.0"
    )
    
    # ============ 样式 ============
    styles = getSampleStyleSheet()
    
    styles.add(ParagraphStyle(
        'CNTitle',
        fontName='STSong-Light',
        fontSize=22,
        leading=30,
        textColor=white,
        alignment=TA_CENTER,
        spaceAfter=0
    ))
    
    styles.add(ParagraphStyle(
        'CNSubtitle',
        fontName='STSong-Light',
        fontSize=12,
        leading=18,
        textColor=white,
        alignment=TA_CENTER,
        spaceAfter=0
    ))
    
    styles.add(ParagraphStyle(
        'CNSection',
        fontName='STSong-Light',
        fontSize=14,
        leading=22,
        textColor=PRIMARY,
        spaceBefore=12,
        spaceAfter=6,
    ))
    
    styles.add(ParagraphStyle(
        'CNLabel',
        fontName='STSong-Light',
        fontSize=11,
        leading=16,
        textColor=TEXT_DARK,
        alignment=TA_LEFT,
    ))
    
    styles.add(ParagraphStyle(
        'CNValue',
        fontName='STSong-Light',
        fontSize=11,
        leading=16,
        textColor=TEXT_DARK,
        alignment=TA_LEFT,
    ))
    
    styles.add(ParagraphStyle(
        'CNFooter',
        fontName='STSong-Light',
        fontSize=8,
        leading=12,
        textColor=TEXT_GRAY,
        alignment=TA_CENTER,
        spaceBefore=4,
    ))
    
    styles.add(ParagraphStyle(
        'CNCenter',
        fontName='STSong-Light',
        fontSize=10,
        leading=14,
        textColor=TEXT_DARK,
        alignment=TA_CENTER,
    ))
    
    # ============ 构建内容 ============
    elements = []
    
    # ---- 标题横幅 ----
    # 用表格模拟彩色横幅
    banner_data = [
        [Paragraph('  智能预测报告', styles['CNTitle'])],
        [Paragraph('  工程开标数据智能分析平台 V6.0', styles['CNSubtitle'])],
    ]
    banner_table = Table(banner_data, colWidths=[A4[0] - 40*mm])
    banner_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, -1), white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(banner_table)
    
    # 导出时间
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(f'导出时间：{export_time}', ParagraphStyle(
        'CNTime', fontName='STSong-Light', fontSize=9, textColor=TEXT_GRAY, alignment=TA_CENTER,
    )))
    elements.append(Spacer(1, 4*mm))
    
    # ---- 项目基本信息 ----
    elements.append(Paragraph('  项目基本信息', styles['CNSection']))
    
    info_fields = [
        ('项目名称', record.get('project_name') or '未命名'),
        ('开标地点', record.get('location_filter') or '不限'),
        ('方法类别', f'方法{record.get("method_filter") or "不限"}'),
        ('日期范围', f'{record.get("date_from") or "—"} ~ {record.get("date_to") or "—"}'),
        ('数据条数', f'{record.get("data_count", 0)} 条'),
        ('AI 辅助', '是' if record.get('used_ai') else '否'),
        ('预测时间', record.get('prediction_time') or '—'),
    ]
    
    info_data = []
    for label, value in info_fields:
        info_data.append([
            Paragraph(f'  {label}：', ParagraphStyle(
                'CNInfoLabel', fontName='STSong-Light', fontSize=11, textColor=TEXT_DARK,
            )),
            Paragraph(value, ParagraphStyle(
                'CNInfoValue', fontName='STSong-Light', fontSize=11, bold=True, textColor=TEXT_DARK,
            )),
        ])
    
    info_table = Table(info_data, colWidths=[80*mm, A4[0]-120*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), BG_LIGHT),
        ('BACKGROUND', (1, 0), (1, -1), white),
        ('TEXTCOLOR', (0, 0), (-1, -1), TEXT_DARK),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
    ]))
    elements.append(info_table)
    
    # ---- 预测结果摘要 ----
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph('  预测结果摘要', styles['CNSection']))
    
    def get_confidence_info(confidence):
        if confidence is None:
            return "未评估", CONF_NONE_BG, TEXT_GRAY
        if confidence >= 0.8:
            return "极高", CONF_HIGH_BG, ACCENT_GREEN
        if confidence >= 0.7:
            return "高", CONF_HIGH_BG, ACCENT_GREEN
        if confidence >= 0.5:
            return "中等", CONF_MED_BG, ACCENT_YELLOW
        return "低", CONF_LOW_BG, ACCENT_RED
    
    pred_data = [[
        Paragraph('  预测参数', ParagraphStyle('CNPredHead', fontName='STSong-Light', fontSize=11, bold=True, textColor=white)),
        Paragraph('  预测值', ParagraphStyle('CNPredHead', fontName='STSong-Light', fontSize=11, bold=True, textColor=white)),
        Paragraph('  置信度', ParagraphStyle('CNPredHead', fontName='STSong-Light', fontSize=11, bold=True, textColor=white)),
        Paragraph('  预测方法', ParagraphStyle('CNPredHead', fontName='STSong-Light', fontSize=11, bold=True, textColor=white)),
    ]]
    
    for param, value, confidence, method in [
        ('方法类别', record.get('method_prediction') or '—', record.get('method_confidence'), record.get('method_prediction') or '—'),
        ('K1 值', record.get('k1_prediction') or '—', record.get('k1_confidence'), record.get('k1_method') or '—'),
        ('Q1 值', record.get('q1_prediction') or '—', record.get('q1_confidence'), record.get('q1_method') or '—'),
    ]:
        conf_text, conf_bg, conf_color = get_confidence_info(confidence)
        pred_data.append([
            Paragraph(f'  {param}', ParagraphStyle('CNPredLabel', fontName='STSong-Light', fontSize=11, textColor=TEXT_DARK)),
            Paragraph(f'  {value}', ParagraphStyle('CNPredVal', fontName='STSong-Light', fontSize=13, bold=True, textColor=PRIMARY)),
            Paragraph(f'  {conf_text} ({confidence:.1%})' if confidence else '  未评估',
                      ParagraphStyle('CNPredConf', fontName='STSong-Light', fontSize=10, bold=True, textColor=conf_color)),
            Paragraph(f'  {method}', ParagraphStyle('CNPredMethod', fontName='STSong-Light', fontSize=10, textColor=TEXT_GRAY)),
        ])
    
    pred_table = Table(pred_data, colWidths=[45*mm, 40*mm, 35*mm, 55*mm])
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
    ]
    # 置信度背景色
    for i in range(1, 4):
        conf_text, conf_bg, conf_color = get_confidence_info(
            [record.get('method_confidence'), record.get('k1_confidence'), record.get('q1_confidence')][i-1]
        )
        table_styles.append(('BACKGROUND', (2, i), (2, i), conf_bg))
        # 预测值行交替底色
        table_styles.append(('BACKGROUND', (0, i), (1, i), BG_LIGHT if i % 2 == 1 else white))
        table_styles.append(('BACKGROUND', (3, i), (3, i), BG_LIGHT if i % 2 == 1 else white))
    
    pred_table.setStyle(TableStyle(table_styles))
    elements.append(pred_table)
    
    # ============ K1 Top5 算法明细（如果有） ============
    if k1_algorithms and len(k1_algorithms) > 0:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph('  🎯 K1 Top5 算法预测结果', styles['CNSection']))
        
        k1_data = [[
            Paragraph('  排名', ParagraphStyle('CNK1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  算法名称', ParagraphStyle('CNK1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  预测值', ParagraphStyle('CNK1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  置信度', ParagraphStyle('CNK1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
        ]]
        
        for idx, algo in enumerate(k1_algorithms[:5], 1):
            bg = BG_LIGHT if idx % 2 == 1 else white
            confidence = algo.get('confidence', 0)
            conf_text, conf_bg, conf_color = get_confidence_info(confidence)
            
            k1_data.append([
                Paragraph(f'  #{idx}', ParagraphStyle('CNK1Rank', fontName='STSong-Light', fontSize=10, bold=True, textColor=PRIMARY)),
                Paragraph(f'  {algo.get("algorithm_name", "—")}', ParagraphStyle('CNK1Name', fontName='STSong-Light', fontSize=10, textColor=TEXT_DARK)),
                Paragraph(f'  {algo.get("prediction_value", "—")}', ParagraphStyle('CNK1Pred', fontName='STSong-Light', fontSize=10, bold=True, textColor=PRIMARY)),
                Paragraph(f'  {conf_text} ({confidence:.1%})', ParagraphStyle('CNK1Conf', fontName='STSong-Light', fontSize=9, bold=True, textColor=conf_color)),
            ])
        
        k1_table = Table(k1_data, colWidths=[20*mm, 70*mm, 35*mm, 45*mm])
        k1_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), ACCENT_GREEN),
            ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
        ]
        # 置信度列背景色
        for i in range(1, len(k1_data)):
            confidence = k1_algorithms[i-1].get('confidence', 0)
            _, conf_bg, _ = get_confidence_info(confidence)
            k1_styles.append(('BACKGROUND', (3, i), (3, i), conf_bg))
        # 数据行交替底色
        for i in range(1, len(k1_data)):
            bg = BG_LIGHT if i % 2 == 1 else white
            k1_styles.append(('BACKGROUND', (0, i), (2, i), bg))
        
        k1_table.setStyle(TableStyle(k1_styles))
        elements.append(k1_table)
    
    # ============ Q1 Top5 算法明细（如果有且非方法1） ============
    if q1_algorithms and len(q1_algorithms) > 0 and record.get('method_prediction') != '1':
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph('  ⚙️ Q1 Top5 算法预测结果', styles['CNSection']))
        
        q1_data = [[
            Paragraph('  排名', ParagraphStyle('CNQ1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  算法名称', ParagraphStyle('CNQ1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  预测值', ParagraphStyle('CNQ1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
            Paragraph('  置信度', ParagraphStyle('CNQ1Head', fontName='STSong-Light', fontSize=10, bold=True, textColor=white)),
        ]]
        
        for idx, algo in enumerate(q1_algorithms[:5], 1):
            bg = BG_LIGHT if idx % 2 == 1 else white
            confidence = algo.get('confidence', 0)
            conf_text, conf_bg, conf_color = get_confidence_info(confidence)
            
            q1_data.append([
                Paragraph(f'  #{idx}', ParagraphStyle('CNQ1Rank', fontName='STSong-Light', fontSize=10, bold=True, textColor=PRIMARY)),
                Paragraph(f'  {algo.get("algorithm_name", "—")}', ParagraphStyle('CNQ1Name', fontName='STSong-Light', fontSize=10, textColor=TEXT_DARK)),
                Paragraph(f'  {algo.get("prediction_value", "—")}', ParagraphStyle('CNQ1Pred', fontName='STSong-Light', fontSize=10, bold=True, textColor=PRIMARY)),
                Paragraph(f'  {conf_text} ({confidence:.1%})', ParagraphStyle('CNQ1Conf', fontName='STSong-Light', fontSize=9, bold=True, textColor=conf_color)),
            ])
        
        q1_table = Table(q1_data, colWidths=[20*mm, 70*mm, 35*mm, 45*mm])
        q1_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), ACCENT_BLUE),
            ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
        ]
        # 置信度列背景色
        for i in range(1, len(q1_data)):
            confidence = q1_algorithms[i-1].get('confidence', 0)
            _, conf_bg, _ = get_confidence_info(confidence)
            q1_styles.append(('BACKGROUND', (3, i), (3, i), conf_bg))
        # 数据行交替底色
        for i in range(1, len(q1_data)):
            bg = BG_LIGHT if i % 2 == 1 else white
            q1_styles.append(('BACKGROUND', (0, i), (2, i), bg))
        
        q1_table.setStyle(TableStyle(q1_styles))
        elements.append(q1_table)
    
    # ---- 底部说明 ----
    elements.append(Spacer(1, 12*mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=BORDER_COLOR))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph('本报告由工程开标数据智能分析平台 V6.0 自动生成', styles['CNFooter']))
    elements.append(Paragraph('预测结果仅供参考，不构成任何决策建议', styles['CNFooter']))
    
    # ============ 构建 PDF ============
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# 辅助函数：获取算法明细
# ============================================================

def get_prediction_details(record_id: int, user_id: int) -> tuple:
    """
    获取预测记录及其Top5算法详情
    Returns: (record_dict, k1_algorithms, q1_algorithms)
    """
    from backend.models.prediction import Prediction
    
    pred_model = Prediction(Config.DATABASE_PATH)
    record = pred_model.get_prediction_by_id(record_id, user_id)
    
    if not record:
        return None, None, None
    
    # 获取Top5算法详情
    k1_algorithms = pred_model.get_algorithm_details(record_id, 'K1')
    q1_algorithms = pred_model.get_algorithm_details(record_id, 'Q1')
    
    return record, k1_algorithms, q1_algorithms
