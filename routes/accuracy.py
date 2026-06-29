#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
预测准确率分析路由模块
处理预测值与真实值的对比分析
"""

from flask import Blueprint, render_template, session, jsonify, request, send_file
from backend.models.prediction import Prediction
from backend.models.operation_log import OperationLog
from backend.config.config import Config
from routes.auth import login_required
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

accuracy_bp = Blueprint('accuracy', __name__)

# 初始化模型
db_path = Config.DATABASE_PATH
prediction_model = Prediction(db_path)
log_model = OperationLog()


@accuracy_bp.route('/accuracy')
@login_required
def accuracy_page():
    """预测准确率分析页面"""
    return render_template('accuracy.html')


@accuracy_bp.route('/api/accuracy/records')
@login_required
def api_accuracy_records():
    """获取准确率对比记录"""
    user_id = session.get('user_id')
    page = int(request.args.get('page', 1))
    page_size = min(int(request.args.get('page_size', 20)), 50)  # 最大50条
    status = request.args.get('status', '').strip() or None
    keyword = request.args.get('keyword', '').strip() or None

    records, total = prediction_model.get_accuracy_records(
        user_id,
        limit=page_size,
        offset=(page - 1) * page_size,
        status=status,
        keyword=keyword
    )

    return jsonify({
        'success': True,
        'total': total,
        'page': page,
        'page_size': page_size,
        'records': records
    })


@accuracy_bp.route('/api/accuracy/auto-match', methods=['POST'])
@login_required
def api_auto_match():
    """自动匹配真实值 - 从开标数据中按项目名称匹配回填"""
    user_id = session.get('user_id')
    result = prediction_model.auto_match_actual_values(user_id)
    
    # 记录操作日志
    if result.get('success'):
        log_model.log(
            user_id=user_id,
            username=session.get('username', 'unknown'),
            action='AUTO_MATCH',
            module='预测准确率',
            table_name='prediction_records',
            new_value={
                'matched_count': result.get('matched_count', 0),
                'skipped_count': result.get('skipped_count', 0),
                'total_predictions': result.get('total_predictions', 0)
            },
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', ''),
            status='success'
        )
    else:
        log_model.log(
            user_id=user_id,
            username=session.get('username', 'unknown'),
            action='AUTO_MATCH',
            module='预测准确率',
            table_name='prediction_records',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', ''),
            status='failed',
            error_message=result.get('error', '未知错误')
        )
    
    return jsonify(result)


@accuracy_bp.route('/api/accuracy/update', methods=['POST'])
@login_required
def api_update_actual():
    """更新真实值"""
    user_id = session.get('user_id')
    data = request.get_json()
    record_id = data.get('id')

    if not record_id:
        return jsonify({'success': False, 'error': '记录 ID 不能为空'})

    success = prediction_model.update_actual_values(
        record_id,
        user_id,
        k1_actual=data.get('k1_actual'),
        q1_actual=data.get('q1_actual'),
        method_actual=data.get('method_actual')
    )

    if success:
        return jsonify({'success': True, 'message': '更新成功'})
    else:
        return jsonify({'success': False, 'error': '更新失败或记录不存在'})


@accuracy_bp.route('/api/accuracy/stats')
@login_required
def api_accuracy_stats():
    """获取准确率统计信息"""
    user_id = session.get('user_id')
    conn = prediction_model.db_path
    import sqlite3

    db = sqlite3.connect(conn)
    cursor = db.cursor()

    # 总预测数
    cursor.execute('SELECT COUNT(*) FROM prediction_records WHERE user_id = ?', (user_id,))
    total = cursor.fetchone()[0]

    # 检查新字段是否存在
    cursor.execute('PRAGMA table_info(prediction_records)')
    columns = [row[1] for row in cursor.fetchall()]
    has_accuracy = 'accuracy_checked' in columns
    has_k1_actual = 'k1_actual' in columns
    has_q1_actual = 'q1_actual' in columns
    has_method_actual = 'method_actual' in columns

    # 已对比数
    if has_accuracy:
        cursor.execute('SELECT COUNT(*) FROM prediction_records WHERE user_id = ? AND accuracy_checked = 1', (user_id,))
        checked = cursor.fetchone()[0]
    else:
        checked = 0

    # 方法类别正确数
    method_correct = 0
    method_total = 0
    if has_accuracy and has_method_actual:
        cursor.execute('''
            SELECT COUNT(*) FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND method_prediction IS NOT NULL AND method_actual IS NOT NULL
            AND method_prediction = method_actual
        ''', (user_id,))
        method_correct = cursor.fetchone()[0]

        cursor.execute('''
            SELECT COUNT(*) FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND method_prediction IS NOT NULL AND method_actual IS NOT NULL
        ''', (user_id,))
        method_total = cursor.fetchone()[0]

    # K1 偏差分布
    k1_rows = []
    if has_accuracy and has_k1_actual:
        cursor.execute('''
            SELECT k1_prediction, k1_actual FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND k1_prediction IS NOT NULL AND k1_actual IS NOT NULL
        ''', (user_id,))
        k1_rows = cursor.fetchall()

    k1_accurate = 0
    k1_normal = 0
    k1_deviated = 0
    k1_total = len(k1_rows)

    for pred, actual in k1_rows:
        try:
            dev = abs(float(pred) - float(actual))
            if dev == 0:
                k1_accurate += 1
            elif dev <= 0.01:
                k1_normal += 1
            else:
                k1_deviated += 1
        except (ValueError, TypeError):
            pass

    # Q1 偏差分布
    q1_rows = []
    if has_accuracy and has_q1_actual:
        cursor.execute('''
            SELECT q1_prediction, q1_actual FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND q1_prediction IS NOT NULL AND q1_actual IS NOT NULL
        ''', (user_id,))
        q1_rows = cursor.fetchall()

    q1_accurate = 0
    q1_normal = 0
    q1_deviated = 0
    q1_total = len(q1_rows)

    for pred, actual in q1_rows:
        try:
            dev = abs(float(pred) - float(actual))
            if dev == 0:
                q1_accurate += 1
            elif dev <= 0.1:
                q1_normal += 1
            else:
                q1_deviated += 1
        except (ValueError, TypeError):
            pass

    # Q1 精准率（排除方法类别为1的记录）
    q1_excl_rows = []
    if has_accuracy and has_q1_actual:
        cursor.execute('''
            SELECT q1_prediction, q1_actual FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND q1_prediction IS NOT NULL AND q1_actual IS NOT NULL
            AND (method_prediction IS NULL OR method_prediction != '1')
        ''', (user_id,))
        q1_excl_rows = cursor.fetchall()

    q1_excl_accurate = 0
    q1_excl_total = len(q1_excl_rows)

    for pred, actual in q1_excl_rows:
        try:
            dev = abs(float(pred) - float(actual))
            if dev == 0:
                q1_excl_accurate += 1
        except (ValueError, TypeError):
            pass

    db.close()

    method_rate = round(method_correct / method_total * 100, 1) if method_total > 0 else 0
    q1_excl_rate = round(q1_excl_accurate / q1_excl_total * 100, 1) if q1_excl_total > 0 else 0

    return jsonify({
        'success': True,
        'stats': {
            'total': total,
            'checked': checked,
            'method': {
                'correct': method_correct,
                'total': method_total,
                'rate': method_rate
            },
            'k1': {
                'accurate': k1_accurate,
                'normal': k1_normal,
                'deviated': k1_deviated,
                'total': k1_total
            },
            'q1': {
                'accurate': q1_accurate,
                'normal': q1_normal,
                'deviated': q1_deviated,
                'total': q1_total
            },
            'q1_excl': {
                'accurate': q1_excl_accurate,
                'total': q1_excl_total,
                'rate': q1_excl_rate
            }
        }
    })


@accuracy_bp.route('/api/accuracy/trend')
@login_required
def api_accuracy_trend():
    """获取准确率趋势数据（按月）"""
    user_id = session.get('user_id')
    conn = prediction_model.db_path
    import sqlite3

    db = sqlite3.connect(conn)
    cursor = db.cursor()

    # 检查字段是否存在
    cursor.execute('PRAGMA table_info(prediction_records)')
    columns = [row[1] for row in cursor.fetchall()]
    has_k1_actual = 'k1_actual' in columns
    has_q1_actual = 'q1_actual' in columns
    has_method_actual = 'method_actual' in columns

    # 查询最近12个月的已对比记录
    # K1 月度数据
    k1_months = []
    k1_accurate_rates = []
    k1_counts = []
    
    if has_k1_actual:
        cursor.execute('''
            SELECT strftime('%Y-%m', checked_time) as month,
                   k1_prediction, k1_actual
            FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND k1_prediction IS NOT NULL AND k1_actual IS NOT NULL
            AND checked_time IS NOT NULL
            ORDER BY month ASC
        ''', (user_id,))
        
        monthly_data = {}
        for month, pred, actual in cursor.fetchall():
            if month not in monthly_data:
                monthly_data[month] = {'total': 0, 'accurate': 0}
            try:
                dev = abs(float(pred) - float(actual))
                monthly_data[month]['total'] += 1
                if dev == 0:
                    monthly_data[month]['accurate'] += 1
            except (ValueError, TypeError):
                pass
        
        # 只取最近12个月
        sorted_months = sorted(monthly_data.keys())[-12:]
        for month in sorted_months:
            data = monthly_data[month]
            rate = round(data['accurate'] / data['total'] * 100, 1) if data['total'] > 0 else 0
            k1_months.append(month)
            k1_accurate_rates.append(rate)
            k1_counts.append(data['total'])

    # Q1 月度数据（排除方法1）
    q1_months = []
    q1_accurate_rates = []
    q1_counts = []
    
    if has_q1_actual:
        cursor.execute('''
            SELECT strftime('%Y-%m', checked_time) as month,
                   q1_prediction, q1_actual, method_prediction
            FROM prediction_records
            WHERE user_id = ? AND accuracy_checked = 1
            AND q1_prediction IS NOT NULL AND q1_actual IS NOT NULL
            AND (method_prediction IS NULL OR method_prediction != '1')
            AND checked_time IS NOT NULL
            ORDER BY month ASC
        ''', (user_id,))
        
        monthly_data = {}
        for month, pred, actual, method in cursor.fetchall():
            if month not in monthly_data:
                monthly_data[month] = {'total': 0, 'accurate': 0}
            try:
                dev = abs(float(pred) - float(actual))
                monthly_data[month]['total'] += 1
                if dev == 0:
                    monthly_data[month]['accurate'] += 1
            except (ValueError, TypeError):
                pass
        
        sorted_months = sorted(monthly_data.keys())[-12:]
        for month in sorted_months:
            data = monthly_data[month]
            rate = round(data['accurate'] / data['total'] * 100, 1) if data['total'] > 0 else 0
            q1_months.append(month)
            q1_accurate_rates.append(rate)
            q1_counts.append(data['total'])

    db.close()

    return jsonify({
        'success': True,
        'trend': {
            'k1': {
                'months': k1_months,
                'accurate_rates': k1_accurate_rates,
                'counts': k1_counts
            },
            'q1': {
                'months': q1_months,
                'accurate_rates': q1_accurate_rates,
                'counts': q1_counts
            }
        }
    })


@accuracy_bp.route('/api/accuracy/timeline')
@login_required
def api_accuracy_timeline():
    """获取预测值与真实值的时间序列对比数据"""
    user_id = session.get('user_id')
    conn = prediction_model.db_path
    import sqlite3

    db = sqlite3.connect(conn)
    cursor = db.cursor()

    # 检查字段是否存在
    cursor.execute('PRAGMA table_info(prediction_records)')
    columns = [row[1] for row in cursor.fetchall()]
    has_k1_actual = 'k1_actual' in columns
    has_q1_actual = 'q1_actual' in columns
    has_method_actual = 'method_actual' in columns

    # 查询所有已录入真实值的记录，按预测时间排序
    cursor.execute('''
        SELECT id, project_name, prediction_time,
               method_prediction, method_actual,
               k1_prediction, k1_actual,
               q1_prediction, q1_actual
        FROM prediction_records
        WHERE user_id = ? AND accuracy_checked = 1
        ORDER BY prediction_time ASC
    ''', (user_id,))

    records = []
    for row in cursor.fetchall():
        rid, project_name, pred_time, method_pred, method_act, k1_pred, k1_act, q1_pred, q1_act = row
        item = {
            'id': rid,
            'project_name': project_name or '',
            'time': str(pred_time)[:16] if pred_time else '',
        }

        # 方法类别（转为数值：方法1=1, 方法2=2）
        if method_pred and method_act:
            item['method_pred'] = int(method_pred)
            item['method_act'] = int(method_act)
        else:
            item['method_pred'] = None
            item['method_act'] = None

        # K1
        if k1_pred is not None and k1_act is not None:
            try:
                item['k1_pred'] = round(float(k1_pred), 4)
                item['k1_act'] = round(float(k1_act), 4)
            except (ValueError, TypeError):
                item['k1_pred'] = None
                item['k1_act'] = None
        else:
            item['k1_pred'] = None
            item['k1_act'] = None

        # Q1
        if q1_pred is not None and q1_act is not None:
            try:
                item['q1_pred'] = round(float(q1_pred), 4)
                item['q1_act'] = round(float(q1_act), 4)
            except (ValueError, TypeError):
                item['q1_pred'] = None
                item['q1_act'] = None
        else:
            item['q1_pred'] = None
            item['q1_act'] = None

        records.append(item)

    db.close()

    return jsonify({
        'success': True,
        'timeline': records
    })


@accuracy_bp.route('/api/accuracy/export')
@login_required
def api_export_accuracy():
    """导出预测准确率为 Excel 文件"""
    user_id = session.get('user_id')
    keyword = request.args.get('keyword', '').strip() or None
    status = request.args.get('status', '').strip() or None

    # 获取所有记录（不分页）
    records, total = prediction_model.get_accuracy_records(
        user_id,
        limit=10000,
        offset=0,
        status=status,
        keyword=keyword
    )

    if not records:
        return jsonify({'success': False, 'error': '没有可导出的数据'})

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '预测准确率分析'

    # 样式定义
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 高亮填充
    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 表头
    headers = [
        '序号', '项目名称', '预测时间', '方法类别\n(预测)', '方法类别\n(真实)', '方法\n是否正确',
        'K1 预测值', 'K1 真实值', 'K1 偏差', 'K1 偏差等级',
        'Q1 预测值', 'Q1 真实值', 'Q1 偏差', 'Q1 偏差等级',
        '对比时间'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    level_map = {
        'accurate': '精准',
        'normal': '一般',
        'deviated': '偏差大',
        'unchecked': '未对比'
    }
    fill_map = {
        'accurate': success_fill,
        'normal': warn_fill,
        'deviated': error_fill,
        'unchecked': None
    }

    for idx, record in enumerate(records, 1):
        row = idx + 1
        
        # 方法类别
        method_pred = record.get('method_prediction', '') or ''
        method_pred_display = '方法1' if method_pred == '1' else ('方法2' if method_pred == '2' else '')
        method_actual = record.get('method_actual', '') or ''
        method_actual_display = '方法1' if method_actual == '1' else ('方法2' if method_actual == '2' else '')
        method_correct = '✅ 正确' if record.get('method_correct') else ('❌ 偏差' if record.get('method_correct') is False else '⏳ 待录入')

        # K1 偏差
        k1_dev = record.get('k1_deviation')
        k1_dev_display = round(k1_dev, 4) if k1_dev is not None else ''
        k1_level = record.get('k1_level', 'unchecked')
        k1_level_display = level_map.get(k1_level, '未对比')

        # Q1 偏差
        q1_dev = record.get('q1_deviation')
        q1_dev_display = round(q1_dev, 4) if q1_dev is not None else ''
        q1_level = record.get('q1_level', 'unchecked')
        q1_level_display = level_map.get(q1_level, '未对比')

        # 对比时间
        checked_time = record.get('checked_time', '') or ''
        if checked_time:
            checked_time = str(checked_time)[:19]

        values = [
            idx,
            record.get('project_name', ''),
            str(record.get('prediction_time', ''))[:19] if record.get('prediction_time') else '',
            method_pred_display,
            method_actual_display,
            method_correct,
            record.get('k1_prediction', '') or '',
            record.get('k1_actual', '') or '',
            k1_dev_display,
            k1_level_display,
            record.get('q1_prediction', '') or '',
            record.get('q1_actual', '') or '',
            q1_dev_display,
            q1_level_display,
            checked_time
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

            # K1/Q1 偏差等级列高亮
            if col == 10 and k1_level in fill_map and fill_map[k1_level]:
                cell.fill = fill_map[k1_level]
            if col == 14 and q1_level in fill_map and fill_map[q1_level]:
                cell.fill = fill_map[q1_level]
            # 方法正确列高亮
            if col == 6:
                if '正确' in str(value):
                    cell.fill = success_fill
                elif '偏差' in str(value):
                    cell.fill = error_fill

    # 设置列宽
    col_widths = [6, 30, 18, 12, 12, 12, 12, 12, 10, 12, 12, 12, 10, 12, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 导出文件
    filename = f'预测准确率分析_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 记录操作日志
    log_model.log(
        user_id=user_id,
        username=session.get('username', 'unknown'),
        action='EXPORT',
        module='预测准确率',
        table_name='prediction_records',
        new_value={'record_count': len(records), 'filename': filename},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent', ''),
        status='success'
    )

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
